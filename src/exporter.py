"""
导出模块 — Excel/CSV/JSON
"""
from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

from src.models import MergedTeacherRow, StatsOverview
from src.config import EXPORT_DIR
from src.utils import sanitize_for_excel, setup_logger

logger = setup_logger("exporter")

# 列名映射
COLUMNS = {
    "index": "序号",
    "teacher_name": "导师姓名",
    "institution": "学校/研究单位",
    "teacher_email": "导师邮箱",
    "first_sent_at": "首次发送时间",
    "last_sent_at": "最近发送时间",
    "send_count": "发送次数",
    "correspondence_summary": "邮件往来情况",
    "has_real_reply": "是否收到真实回复",
    "teacher_reply_time": "导师回复时间",
    "actual_responder": "实际回复人",
    "reply_content_summary": "导师回复核心内容",
    "evidence_sentences": "关键证据句",
    "admission_intent": "招生意向判断",
    "my_reply_status": "我是否已回复",
    "action_required": "现在是否需要回复",
    "recommended_next_step": "建议下一步行动",
    "confidence": "识别置信度",
    "needs_review": "是否需要人工复核",
}


def rows_to_dataframe(rows: list[MergedTeacherRow]) -> pd.DataFrame:
    """将合并表转为 DataFrame"""
    data = []
    for row in rows:
        d = {}
        for field, label in COLUMNS.items():
            val = getattr(row, field, "")
            if isinstance(val, bool):
                val = "是" if val else "否"
            elif hasattr(val, "value"):
                val = val.value
            d[label] = sanitize_for_excel(str(val))
        data.append(d)
    return pd.DataFrame(data)


def export_excel(rows: list[MergedTeacherRow], stats: StatsOverview,
                 messages: list[dict] | None = None,
                 review_items: list[dict] | None = None,
                 scan_errors: list[dict] | None = None,
                 filepath: str | None = None) -> str:
    """导出 Excel 文件"""
    if not filepath:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = str(EXPORT_DIR / f"套磁统计_{timestamp}.xlsx")

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Sheet 1: 导师总表
        df = rows_to_dataframe(rows)
        df.to_excel(writer, sheet_name="导师总表", index=False)

        # Sheet 2: 汇总统计
        stats_data = _stats_to_dict(stats)
        df_stats = pd.DataFrame(list(stats_data.items()), columns=["统计项", "数值"])
        df_stats.to_excel(writer, sheet_name="汇总统计", index=False)

        # Sheet 3: 邮件明细
        if messages:
            safe_msgs = []
            for m in messages[:5000]:
                safe_msgs.append({
                    "邮件ID": m.get("id", ""),
                    "文件夹": m.get("folder_id", ""),
                    "日期": m.get("date", ""),
                    "发件人": m.get("from_name", "") or m.get("from_email", ""),
                    "主题": sanitize_for_excel(m.get("subject", "")),
                    "是否发送": "是" if m.get("is_sent_by_me") else "否",
                    "是否自动回复": "是" if m.get("is_auto_reply") else "否",
                    "是否退信": "是" if m.get("is_bounce") else "否",
                })
            pd.DataFrame(safe_msgs).to_excel(writer, sheet_name="邮件明细", index=False)

        # Sheet 4: 待人工复核
        if review_items:
            pd.DataFrame(review_items).to_excel(writer, sheet_name="待人工复核", index=False)

        # Sheet 5: 扫描错误
        if scan_errors:
            pd.DataFrame(scan_errors).to_excel(writer, sheet_name="扫描错误", index=False)

        # Sheet 6: 配置说明
        config_info = [
            {"项目": "软件名称", "说明": "保研套磁统计工具"},
            {"项目": "导出时间", "说明": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
            {"项目": "邮箱访问模式", "说明": "只读（readonly）"},
            {"项目": "安全提示", "说明": "本文件不包含邮箱授权码或API Key"},
            {"项目": "数据说明", "说明": "所有分析结果均基于邮件内容自动生成，建议人工核实"},
        ]
        pd.DataFrame(config_info).to_excel(writer, sheet_name="配置说明", index=False)

    # 格式化 Excel
    _format_excel(filepath)

    logger.info(f"Excel 导出完成: {filepath}")
    return filepath


def export_csv(rows: list[MergedTeacherRow], filepath: str | None = None) -> str:
    """导出 CSV 文件"""
    if not filepath:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = str(EXPORT_DIR / f"套磁统计_{timestamp}.csv")

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)

    df = rows_to_dataframe(rows)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")

    logger.info(f"CSV 导出完成: {filepath}")
    return filepath


def export_json(rows: list[MergedTeacherRow], stats: StatsOverview,
                filepath: str | None = None) -> str:
    """导出 JSON 文件"""
    if not filepath:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = str(EXPORT_DIR / f"套磁统计_{timestamp}.json")

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)

    data = {
        "export_time": datetime.now().isoformat(),
        "stats": _stats_to_dict(stats),
        "teachers": [row.model_dump() for row in rows],
    }

    # 处理枚举值序列化
    for t in data["teachers"]:
        for k, v in t.items():
            if hasattr(v, "value"):
                t[k] = v.value

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    logger.info(f"JSON 导出完成: {filepath}")
    return filepath


def _stats_to_dict(stats: StatsOverview) -> dict:
    """统计数据转字典"""
    return {
        "相关邮件总数": stats.total_related_emails,
        "已发送相关邮件数": stats.sent_emails,
        "有效回复邮件数": stats.received_reply_emails,
        "自动回复数": stats.auto_reply_count,
        "退信数": stats.bounce_count,
        "去重后导师总数": stats.teacher_count,
        "已确认收到真实回复人数": stats.replied_teacher_count,
        "无法确认回复人数": stats.uncertain_reply_count,
        "未回复人数": stats.no_reply_count,
        "有效回复率(%)": stats.reply_rate,
        "明确积极意向人数": stats.positive_intent_count,
        "欢迎报名但仍需考核人数": stats.welcome_apply_count,
        "建议参加夏令营/预推免人数": stats.suggest_camp_count,
        "愿意进一步交流人数": stats.willing_to_talk_count,
        "暂时无法确定名额人数": stats.quota_uncertain_count,
        "礼貌回复态度不明确人数": stats.polite_no_intent_count,
        "名额已满/拒绝人数": stats.negative_count,
        "需要立即回复人数": stats.need_reply_now_count,
        "建议跟进人数": stats.suggest_followup_count,
        "需要人工复核人数": stats.needs_review_count,
        "数据一致性": "通过" if stats.consistency_valid else stats.consistency_message,
    }


def _format_excel(filepath: str) -> None:
    """格式化 Excel 文件"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath)

        for ws in wb.worksheets:
            # 冻结首行
            ws.freeze_panes = "A2"

            # 自动筛选
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions

            # 表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 调整列宽
            for col in range(1, ws.max_column + 1):
                max_len = 0
                for row in range(1, min(ws.max_row + 1, 50)):
                    val = str(ws.cell(row=row, column=col).value or "")
                    max_len = max(max_len, len(val))
                width = min(max(max_len * 1.5 + 2, 10), 50)
                ws.column_dimensions[get_column_letter(col)].width = width

            # 内容自动换行
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).alignment = Alignment(
                        wrap_text=True, vertical="top")

        wb.save(filepath)
    except Exception as e:
        logger.warning(f"Excel 格式化失败: {e}")
